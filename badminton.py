from csv import excel
from unittest import result
from webbrowser import BaseBrowser
from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import numpy as np
from openpyxl import Workbook, load_workbook
import re
from HelpFunctions import *
import numpy as np
from ranking import navne
PATH = "C:\webdrivers\chromedriver.exe" 
driver = webdriver.Chrome(PATH)
driver.get(link)
main = driver.find_element_by_class_name("container__fsbody")
tekst = str(main.text)
navn = navne
def stats(x,name): 
    Liste,book,navneliste,længde,top = kategori(name)
    loop = 0
    rank = navne
    if name == "dd" or "hd" or "md":
        loop = 30
    else: 
        loop = 20
    pattern = "[0-9]+"
    wb = load_workbook(book)
    ws = wb.active   
    for i in range(len(x)): 
        
        if i == len(x): 
                    break
        for j in range(len(Liste)): 

            if x[i:i+(navneliste[j])] == Liste[j]: 
                sejr = False
                femPlus = False
                femMinus = False
                dårligere = False
                bedre = False
                bedrePlus = False
                null = False
                dårligerePlus = False
                excelNummer = people(Liste[j],Liste)+2   
                modstandergennemsnitspoint = 0
                gennemsnitspoint = 0
                antagonistpoint = []
                modstanderpoint = []
                længde = 0
                egenrank = 0
                modstanderrank = 0
                trukket = (x[i:i+navneliste[j]+20])
                kamp = False
                if "-" in trukket: 
                    #ws['AA'+str(excelNummer)].value = ws['AA'+str(excelNummer)].value+0.5
                    #print(Liste[j],excelNummer,"spiller trukket sig")
                    #print(Liste[j])
                    kamp = True
                    break
                elif "Aflyst" in trukket: 
                    #ws['AA'+str(excelNummer)].value = ws['AA'+str(excelNummer)].value+0.5
                    #print("kamp er aflyst")
                    #print(Liste[j])
                    kamp = True
                    break
                elif "(trukket sig)" in trukket: 
                    #ws['AA'+str(excelNummer)].value = ws['AA'+str(excelNummer)].value+0.5
                    #print("tror der er blevet trukket en spiller")
                    #print(Liste[j])
                    kamp = True
                    break
                elif "WO" in trukket: 
                    kamp = True
                    break
            

                if ((x[i+navneliste[j]+1])).isalpha(): 
                    forrigresultat = x[i:i+loop]
                    resultat = (re.findall(pattern,x[i:i+navneliste[j]+60]))
                    længde = 0
                    if len(resultat) > 5:
                        
                        if int(resultat[0])+int(resultat[1]) == 3 :  
                            længde = 8
                        elif int(resultat[0])+int(resultat[1]) == 2:
                            længde = 6
                       
                        if int(resultat[0])+int(resultat[1]) < 1: 
                                break
                        else:     
                            ws['Q'+str(excelNummer)].value = ws['Q'+str(excelNummer)].value+1
                        for s in range(2,længde): 
                                if s%2 == 1: 
                                    modstanderpoint.append(int(resultat[s]))
                                else: 
                                    antagonistpoint.append(int(resultat[s]))
                        if len(antagonistpoint) != len(modstanderpoint): 
                                antagonistpoint.pop()
                        if len(modstanderpoint) >2:
                                if (modstanderpoint[0] > antagonistpoint[0]) :
                                    if (modstanderpoint[1] > antagonistpoint[1]):
                                        modstanderpoint.pop() and antagonistpoint.pop()
                                elif (antagonistpoint[0] > modstanderpoint[0]):
                                    if (antagonistpoint[1] > modstanderpoint[1]):
                                        modstanderpoint.pop() and antagonistpoint.pop()
                        
                        gennemsnitspoint = sum(antagonistpoint)
                        modstandergennemsnitspoint = sum(modstanderpoint)
                        if len(modstanderpoint) < 1 or len(antagonistpoint) < 1: 
                            break
                        if (int(modstanderpoint[-1]) < 20) and int(antagonistpoint[-1]) < 20: 
                                modstanderpoint.pop() and antagonistpoint.pop()
                        for l in range(len(top)): 
                            
                                if top[l] in forrigresultat and Liste[j] != top[l]: 
                                    
                                    ws['Z'+str(excelNummer)].value = ws['Z'+str(excelNummer)].value+1

                                    
                                    sejr = True
                        if int(resultat[0])+int(resultat[1]) == 3:  
                           
                                    if int(antagonistpoint[0]) > int(modstanderpoint[0]) and int(antagonistpoint[1]) < int(modstanderpoint[1]) and int(antagonistpoint[2]) > int(modstanderpoint[2]):
                                        
                                        ws['CO'+str(excelNummer)].value = ws['CO'+str(excelNummer)].value+1
                                    elif int(antagonistpoint[0]) < int(modstanderpoint[0]) and int(antagonistpoint[1]) > int(modstanderpoint[1]) and int(antagonistpoint[2]) > int(modstanderpoint[2]):
                                        ws['CL'+str(excelNummer)].value = ws['CL'+str(excelNummer)].value+1
                                    elif int(antagonistpoint[0]) < int(modstanderpoint[0]) and int(antagonistpoint[1]) > int(modstanderpoint[1]) and int(antagonistpoint[2]) < int(modstanderpoint[2]): 
                                        ws['CM'+str(excelNummer)].value = ws['CM'+str(excelNummer)].value+1
                                    elif int(antagonistpoint[0]) > int(modstanderpoint[0]) and int(antagonistpoint[1]) < int(modstanderpoint[1]) and int(antagonistpoint[2]) < int(modstanderpoint[2]): 
                                        ws['CP'+str(excelNummer)].value = ws['CP'+str(excelNummer)].value+1
                        for h in range(len(rank)):
                                if rank[h][2] in forrigresultat and Liste[j] != rank[h][2]:
                                    modstanderrank = navne[h][0]
                                elif rank[h][2] == Liste[j] : 
                                    egenrank = rank[h][0]
                                
                        if int(egenrank) > int(modstanderrank): 
                            if int(modstanderrank) == 0 or int(egenrank) == 0:
                                null = True
                            if int(modstanderrank) != 0 and int(egenrank) != 0: 
                                dårligere = True
                                if int(egenrank)-int(modstanderrank) > 5 : 
                                    dårligerePlus = True
                                elif int(egenrank)-int(modstanderrank) < 5:
                                    femMinus = True
                        if int(modstanderrank) > int(egenrank): 
                            if int(modstanderrank) == 0 or int(egenrank) == 0:
                                null = True
                            if int(modstanderrank) != 0 and int(egenrank) != 0: 
                                bedre = True
                                if int(modstanderrank)-int(egenrank) > 5 : 
                                    bedrePlus = True
                                elif int(modstanderrank)-int(egenrank) < 5:
                                    femPlus = True
                        
                        if kamp == True: 
                            print("SHIT")
                        ws['R'+str(excelNummer)].value = ws['R'+str(excelNummer)].value+gennemsnitspoint
                        ws['S'+str(excelNummer)].value = ws['S'+str(excelNummer)].value+modstandergennemsnitspoint
                        if sejr == True: 
                            ws['AL'+str(excelNummer)].value = ws['AL'+str(excelNummer)].value+gennemsnitspoint
                            ws['AM'+str(excelNummer)].value = ws['AM'+str(excelNummer)].value+modstandergennemsnitspoint
                        if int(resultat[0]) == 2: 
                            if dårligerePlus == True and null == False: 
                                    ws['CA'+str(excelNummer)].value = ws['CA'+str(excelNummer)].value+1
                            if bedrePlus == True and null == False: 
                                    ws['BP'+str(excelNummer)].value = ws['BP'+str(excelNummer)].value+1
                            if sejr == True:
                                ws['AA'+str(excelNummer)].value = ws['AA'+str(excelNummer)].value+1
                            if femPlus == True and null == False: 
                                ws['AT'+str(excelNummer)].value = ws['AT'+str(excelNummer)].value+1
                            if femMinus == True and null == False: 
                                ws['BE'+str(excelNummer)].value = ws['BE'+str(excelNummer)].value+1
                            
                            ws['E'+str(excelNummer)].value = ws['E'+str(excelNummer)].value+1
                            ws['J'+str(excelNummer)].value = ws['J'+str(excelNummer)].value+1
                                

                            if int(resultat[1]) == 0: 
                                if dårligerePlus == True and null == False: 
                                    ws['CB'+str(excelNummer)].value = ws['CB'+str(excelNummer)].value+1
                                if femMinus == True and null == False: 
                                    ws['BF'+str(excelNummer)].value = ws['BF'+str(excelNummer)].value+1
                            
                                if bedrePlus == True and null == False: 
                                    ws['BQ'+str(excelNummer)].value = ws['BQ'+str(excelNummer)].value+1
                                if sejr == True: 
                                    ws['AB'+str(excelNummer)].value = ws['AB'+str(excelNummer)].value+1
                                    ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+2
                                if femPlus == True and null == False: 
                                    ws['AU'+str(excelNummer)].value = ws['AU'+str(excelNummer)].value+1
                                
                                ws['F'+str(excelNummer)].value = ws['F'+str(excelNummer)].value+1    
                                ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+2
                            elif int(resultat[1]) == 1: 
                                if dårligerePlus == True and null == False: 
                                    ws['CC'+str(excelNummer)].value = ws['CC'+str(excelNummer)].value+1
                                if femMinus == True and null == False: 
                                    ws['BG'+str(excelNummer)].value = ws['BG'+str(excelNummer)].value+1
                            
                                if bedrePlus == True and null == False: 
                                    ws['BR'+str(excelNummer)].value = ws['BR'+str(excelNummer)].value+1
                                if femPlus == True and null == False:  
                                    ws['AV'+str(excelNummer)].value = ws['AV'+str(excelNummer)].value+1
                                
                                if sejr == True: 
                                    ws['AC'+str(excelNummer)].value = ws['AC'+str(excelNummer)].value+1
                                    ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+3
                                ws['G'+str(excelNummer)].value = ws['G'+str(excelNummer)].value+1
                                ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+3
                            break
                        elif int(resultat[0]) == 1: 
                            if dårligerePlus == True and null == False: 
                                    ws['CF'+str(excelNummer)].value = ws['CF'+str(excelNummer)].value+1
                                    ws['CH'+str(excelNummer)].value = ws['CH'+str(excelNummer)].value+1
                            if femMinus == True and null == False: 
                                ws['BJ'+str(excelNummer)].value = ws['BJ'+str(excelNummer)].value+1
                                ws['BL'+str(excelNummer)].value = ws['BL'+str(excelNummer)].value+1
                            
                            if bedrePlus == True and null == False: 
                                    ws['BU'+str(excelNummer)].value = ws['BU'+str(excelNummer)].value+1
                                    ws['BW'+str(excelNummer)].value = ws['BW'+str(excelNummer)].value+1
                            if femPlus == True and null == False: 
                                ws['AY'+str(excelNummer)].value = ws['AY'+str(excelNummer)].value+1
                                ws['BA'+str(excelNummer)].value = ws['BA'+str(excelNummer)].value+1
                                
                            if sejr == True: 
                                ws['AH'+str(excelNummer)].value = ws['AH'+str(excelNummer)].value+1
                                ws['AF'+str(excelNummer)].value = ws['AF'+str(excelNummer)].value+1
                                ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+3
                            ws['N'+str(excelNummer)].value = ws['N'+str(excelNummer)].value+1
                            ws['J'+str(excelNummer)].value = 0
                            ws['L'+str(excelNummer)].value = ws['L'+str(excelNummer)].value+1
                            ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+3
                            break
                        elif int(resultat[0]) == 0: 
                            if dårligerePlus == True and null == False: 
                                    ws['CF'+str(excelNummer)].value = ws['CF'+str(excelNummer)].value+1
                                    ws['CG'+str(excelNummer)].value = ws['CG'+str(excelNummer)].value+1
                            if femMinus == True and null == False: 
                                ws['BJ'+str(excelNummer)].value = ws['BJ'+str(excelNummer)].value+1
                                ws['BK'+str(excelNummer)].value = ws['BK'+str(excelNummer)].value+1
                            
                            
                            if bedrePlus == True and null == False: 
                                    ws['BV'+str(excelNummer)].value = ws['BV'+str(excelNummer)].value+1
                                    ws['BU'+str(excelNummer)].value = ws['BU'+str(excelNummer)].value+1
                            if femPlus == True and null == False: 
                                ws['AY'+str(excelNummer)].value = ws['AY'+str(excelNummer)].value+1
                                ws['AV'+str(excelNummer)].value = ws['AV'+str(excelNummer)].value+1
                                
                                
                            if sejr == True: 
                                ws['AG'+str(excelNummer)].value = ws['AG'+str(excelNummer)].value+1
                                ws['AF'+str(excelNummer)].value = ws['AF'+str(excelNummer)].value+1
                                ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+2
                            ws['J'+str(excelNummer)].value = 0
                            ws['M'+str(excelNummer)].value = ws['M'+str(excelNummer)].value+1
                            ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+2
                            ws['L'+str(excelNummer)].value = ws['L'+str(excelNummer)].value+1
                            break
                    else:
                        break  
                        
                else:
                    resultat = (re.findall(pattern,x[i:i+navneliste[j]+25]))
                    forrigresultat = x[i-loop:i]
                    
                            
                    if len(resultat) > 5:
                        
                        if len(resultat) > 8:  
                            længde = 8
                        else:
                            længde = len(resultat)
                        
                        if int(resultat[0])+int(resultat[1]) < 1: 
                            break
                        else: 
                            ws['Q'+str(excelNummer)].value = ws['Q'+str(excelNummer)].value+1
                        for s in range(2,længde): 
                            if s%2 == 0: 
                                modstanderpoint.append(int(resultat[s]))
  
                            else: 
                                antagonistpoint.append(int(resultat[s]))
                        
                        if len(antagonistpoint) != len(modstanderpoint): 
                            modstanderpoint.pop()
                        if len(modstanderpoint) >2:
                            if (modstanderpoint[0] > antagonistpoint[0]) :
                                if (modstanderpoint[1] > antagonistpoint[1]):
                                    modstanderpoint.pop() and antagonistpoint.pop()
                            elif (antagonistpoint[0] > modstanderpoint[0]) and (antagonistpoint[1] > modstanderpoint[1]):
                                modstanderpoint = modstanderpoint[:-1]
                                antagonistpoint = antagonistpoint[:-1]
                        if len(modstanderpoint) < 1 or len(antagonistpoint) < 1: 
                            break
                        if (int(modstanderpoint[-1]) < 20) and int(antagonistpoint[-1]) < 20: 
                                modstanderpoint.pop() and antagonistpoint.pop()
                        gennemsnitspoint = sum(antagonistpoint)
                        modstandergennemsnitspoint = sum(modstanderpoint)
                        for l in range(len(top)): 
                        
                            if top[l] in forrigresultat and top[l] != Liste[j]: 
                                ws['Z'+str(excelNummer)].value = ws['Z'+str(excelNummer)].value+1
                        
                        if int(resultat[0])+int(resultat[1]) == 3:  
                            
                                if int(antagonistpoint[0]) > int(modstanderpoint[0]) and int(antagonistpoint[1]) < int(modstanderpoint[1]) and int(antagonistpoint[2]) > int(modstanderpoint[2]):
                                    ws['CO'+str(excelNummer)].value = ws['CO'+str(excelNummer)].value+1
                                elif int(antagonistpoint[0]) < int(modstanderpoint[0]) and int(antagonistpoint[1]) > int(modstanderpoint[1]) and int(antagonistpoint[2]) > int(modstanderpoint[2]):
                                    ws['CL'+str(excelNummer)].value = ws['CL'+str(excelNummer)].value+1
                                elif int(antagonistpoint[0]) < int(modstanderpoint[0]) and int(antagonistpoint[1]) > int(modstanderpoint[1]) and int(antagonistpoint[2]) < int(modstanderpoint[2]): 
                                    ws['CM'+str(excelNummer)].value = ws['CM'+str(excelNummer)].value+1
                                elif int(antagonistpoint[0]) > int(modstanderpoint[0]) and int(antagonistpoint[1]) < int(modstanderpoint[1]) and int(antagonistpoint[2]) < int(modstanderpoint[2]): 
                                    ws['CP'+str(excelNummer)].value = ws['CP'+str(excelNummer)].value+1
                        for h in range(len(rank)):
                            if rank[h][2] in forrigresultat and Liste[j] != rank[h][2]:
                                modstandernavn = rank[h][2]
                                modstanderrank = navne[h][0]
                            elif rank[h][2] == Liste[j] : 
                                egenrank = rank[h][0]
                        
                        if int(egenrank) > int(modstanderrank): 
                            if int(modstanderrank) == 0 or int(egenrank) == 0:
                                null = True
                            if int(modstanderrank) != 0 and int(egenrank) != 0: 
                                dårligere = True
                                if int(egenrank)-int(modstanderrank) > 5 : 
                                    dårligerePlus = True
                                elif int(egenrank)-int(modstanderrank) < 5:
                                    femMinus = True
                        if int(modstanderrank) > int(egenrank): 
                            if int(modstanderrank) == 0 or int(egenrank) == 0:
                                null = True
                            if int(modstanderrank) != 0 and int(egenrank) != 0: 
                                bedre = True
                                if int(modstanderrank)-int(egenrank) > 5 : 
                                    bedrePlus = True
                                elif int(modstanderrank)-int(egenrank) < 5:
                                    femPlus = True
                    
                        if sejr == True: 
                            ws['AL'+str(excelNummer)].value = ws['AL'+str(excelNummer)].value+gennemsnitspoint
                            ws['AM'+str(excelNummer)].value = ws['AM'+str(excelNummer)].value+modstandergennemsnitspoint
                        ws['R'+str(excelNummer)].value = ws['R'+str(excelNummer)].value+gennemsnitspoint
                        ws['S'+str(excelNummer)].value = ws['S'+str(excelNummer)].value+modstandergennemsnitspoint
                        if int(resultat[1]) == 2: 
                            if femMinus == True and null == False: 
                                ws['BE'+str(excelNummer)].value = ws['BE'+str(excelNummer)].value+1
                            if dårligerePlus == True and null == False: 
                                ws['CA'+str(excelNummer)].value = ws['CA'+str(excelNummer)].value+1
                            if sejr == True : 
                                ws['AA'+str(excelNummer)].value = ws['AA'+str(excelNummer)].value+1
                            if bedrePlus == True and null == False: 
                                ws['BP'+str(excelNummer)].value = ws['BP'+str(excelNummer)].value+1
                            if femPlus == True and null == False: 
                                ws['AT'+str(excelNummer)].value = ws['AT'+str(excelNummer)].value+1
                            
                            ws['E'+str(excelNummer)].value = ws['E'+str(excelNummer)].value+1
                            ws['J'+str(excelNummer)].value = ws['J'+str(excelNummer)].value+1
                            if int(resultat[0]) == 0: 
                                if dårligerePlus == True and null == False: 
                                    ws['CB'+str(excelNummer)].value = ws['CB'+str(excelNummer)].value+1
                                if femMinus == True and null == False: 
                                    ws['BF'+str(excelNummer)].value = ws['BF'+str(excelNummer)].value+1
                        
                                if femPlus == True and null == False: 
                                    ws['AU'+str(excelNummer)].value = ws['AU'+str(excelNummer)].value+1
                                if bedrePlus == True and null == False: 
                                    ws['BQ'+str(excelNummer)].value = ws['BQ'+str(excelNummer)].value+1
                                if sejr == True: 
                                    ws['AB'+str(excelNummer)].value = ws['AB'+str(excelNummer)].value+1
                                    ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+2
                                ws['F'+str(excelNummer)].value = ws['F'+str(excelNummer)].value+1
                                ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+2
                            elif int(resultat[0]) == 1: 
                                if dårligerePlus == True and null == False: 
                                    ws['CC'+str(excelNummer)].value = ws['CC'+str(excelNummer)].value+1
                                if femMinus == True and null == False: 
                                    ws['BG'+str(excelNummer)].value = ws['BG'+str(excelNummer)].value+1
                        
                                if bedrePlus == True and null == False: 
                                    ws['BR'+str(excelNummer)].value = ws['BR'+str(excelNummer)].value+1
                                if femPlus == True and null == False: 
                                    ws['AV'+str(excelNummer)].value = ws['AV'+str(excelNummer)].value+1

                                if sejr == True:
                                    ws['AC'+str(excelNummer)].value = ws['AC'+str(excelNummer)].value+1
                                    ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+3
                                ws['G'+str(excelNummer)].value = ws['G'+str(excelNummer)].value+1
                                ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+3
                            
                            break
                        elif int(resultat[1]) == 1: 
                            if dårligerePlus == True and null == False: 
                                ws['CF'+str(excelNummer)].value = ws['CF'+str(excelNummer)].value+1
                                ws['CH'+str(excelNummer)].value = ws['CH'+str(excelNummer)].value+1
                            if femMinus == True and null == False: 
                                    ws['BJ'+str(excelNummer)].value = ws['BJ'+str(excelNummer)].value+1 
                                    ws['BL'+str(excelNummer)].value = ws['BL'+str(excelNummer)].value+1
                        
                        
                            if femPlus == True and null == False: 
                                ws['AY'+str(excelNummer)].value = ws['AY'+str(excelNummer)].value+1
                                ws['BA'+str(excelNummer)].value = ws['BA'+str(excelNummer)].value+1
                            if bedrePlus == True and null == False: 
                                ws['BU'+str(excelNummer)].value = ws['BU'+str(excelNummer)].value+1
                                ws['BW'+str(excelNummer)].value = ws['BW'+str(excelNummer)].value+1
                            if sejr == True: 
                                ws['AH'+str(excelNummer)].value = ws['AH'+str(excelNummer)].value+1
                                ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+3
                                ws['AF'+str(excelNummer)].value = ws['AF'+str(excelNummer)].value+1
                            ws['L'+str(excelNummer)].value = ws['L'+str(excelNummer)].value+1
                            ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+3
                            ws['J'+str(excelNummer)].value = 0
                            ws['N'+str(excelNummer)].value = ws['N'+str(excelNummer)].value+1
                            break
                        elif int(resultat[1]) == 0:      
                            if dårligerePlus == True and null == False: 
                                ws['CF'+str(excelNummer)].value = ws['CF'+str(excelNummer)].value+1
                                ws['CG'+str(excelNummer)].value = ws['CG'+str(excelNummer)].value+1
                            if femMinus == True and null == False: 
                                    ws['BJ'+str(excelNummer)].value = ws['BJ'+str(excelNummer)].value+1 
                                    ws['BK'+str(excelNummer)].value = ws['BK'+str(excelNummer)].value+1
                        
                         
                            if femPlus == True and null == False: 
                                ws['AY'+str(excelNummer)].value = ws['AY'+str(excelNummer)].value+1
                                ws['AZ'+str(excelNummer)].value = ws['AZ'+str(excelNummer)].value+1
                            if bedrePlus == True and null == False: 
                                ws['BU'+str(excelNummer)].value = ws['BU'+str(excelNummer)].value+1
                                ws['BV'+str(excelNummer)].value = ws['BV'+str(excelNummer)].value+1
                             
                            if sejr == True:
                                ws['AG'+str(excelNummer)].value = ws['AG'+str(excelNummer)].value+1
                                ws['AF'+str(excelNummer)].value = ws['AF'+str(excelNummer)].value+1
                                ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+2
                            ws['M'+str(excelNummer)].value = ws['M'+str(excelNummer)].value+1
                            ws['J'+str(excelNummer)].value =0              
                            ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+2
                            ws['L'+str(excelNummer)].value = ws['L'+str(excelNummer)].value+1

                            break
                        else: 
                            break
                    else: 
                        break
                    
                    

                    
    wb.save(book) 
    return("Excel er opdateret") 

#print(stats(tekst,"hs"))
def rank(x,name) : 
    Liste,book,navneliste,længde,top = kategori(name)
    wb = load_workbook(book)
    ws = wb.active   
    for i in range(len(x)): 
        for j in range(len(Liste)): 
            if x[i][2] == Liste[j]: 
                excelNummer = people(Liste[j],Liste)+2   
                ws['D'+str(excelNummer)].value = x[i][0]
        wb.save(book) 
    return("Ranking er opdateret") 
#print(rank(navne,"hs"))


def loopBadminton(): 
    lst = ["hs","ds","hd","md","dd"]
    tekstlst = [tekst[HSstart:HSslut],tekst[DSstart:DSslut],tekst[HDstart:HDslut],tekst[MDstart:MDslut],tekst[DDstart:DDslut]]
    for i in range(len(lst)):
        stats(tekstlst[i],lst[i])

#print(loopBadminton())
link = ["https://www.flashscore.dk/badminton/bwf-world-tour-maend/french-open/resultater/"]
newlink = ["https://www.flashscore.dk/badminton/bwf-world-tour-maend/all-england-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/chinese-taipei-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/denmark-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/french-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/german-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/hylo-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/india-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/indonesia-masters/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/indonesia-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/korea-masters/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/korea-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/malaysia-masters/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/malaysia-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/odisha-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/orleans-masters/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/singapore-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/spain-masters/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/swiss-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/syed-modi-international-championships/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/thailand-open/resultater/","https://www.flashscore.dk/badminton/bwf-world-tour-maend/world-tour-finals/resultater/","https://www.flashscore.dk/badminton/bwf-maend/vm/"]
def forall(x): 
    
    for i in range(len(x)): 
        print(len(x))
        print(i)
        PATH = "C:\webdrivers\chromedriver.exe" 
        driver = webdriver.Chrome(PATH)
        driver.get(x[i])
        main = driver.find_element_by_class_name("container__fsbody")
        tekst = str(main.text)
        print(stats(tekst,"hs"))

    return("SHIIT")
print(forall(newlink))






driver.quit()