from asyncore import loop
import pandas as pd
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from HelpFunctions import *

def wipe(name): 
    Liste,book,navneliste,længde,top = kategori(name)
    wb = load_workbook(book)
    print(book)
    ws = wb.active 
    for i in range(2,længde+2): 
        ws['E'+str(i)].value = 0
        ws['D'+str(i)].value = 0
        ws['F'+str(i)].value = 0
        ws['M'+str(i)].value = 0
        ws['G'+str(i)].value = 0
        ws['N'+str(i)].value = 0
        ws['O'+str(i)].value = 0
        ws['L'+str(i)].value = 0
        ws['Q'+str(i)].value = 0
        ws['J'+str(i)].value = 0
        ws['R'+str(i)].value = 0
        ws['S'+str(i)].value = 0
        ws['T'+str(i)].value = 0
        ws['X'+str(i)].value = 0
        ws['Z'+str(i)].value = 0
        ws['AA'+str(i)].value = 0
        ws['AB'+str(i)].value = 0
        ws['AC'+str(i)].value = 0
        ws['AF'+str(i)].value = 0
        ws['AG'+str(i)].value = 0
        ws['AG'+str(i)].value = 0
        ws['AH'+str(i)].value = 0
        ws['AL'+str(i)].value = 0
        ws['AM'+str(i)].value = 0
        ws['AN'+str(i)].value = 0
        ws['AT'+str(i)].value = 0
        ws['AU'+str(i)].value = 0
        ws['AV'+str(i)].value = 0
        ws['AY'+str(i)].value = 0
        ws['AZ'+str(i)].value = 0
        ws['BA'+str(i)].value = 0
        ws['BE'+str(i)].value = 0
        ws['BF'+str(i)].value = 0
        ws['BG'+str(i)].value = 0
        ws['BJ'+str(i)].value = 0
        ws['BK'+str(i)].value = 0
        ws['BL'+str(i)].value = 0
        ws['BP'+str(i)].value = 0
        ws['BQ'+str(i)].value = 0
        ws['BR'+str(i)].value = 0
        ws['BU'+str(i)].value = 0
        ws['BV'+str(i)].value = 0
        ws['BW'+str(i)].value = 0
        ws['CA'+str(i)].value = 0
        ws['CB'+str(i)].value = 0
        ws['CC'+str(i)].value = 0
        ws['CF'+str(i)].value = 0
        ws['CG'+str(i)].value = 0
        ws['CH'+str(i)].value = 0
        ws['CL'+str(i)].value = 0
        ws['CM'+str(i)].value = 0
        ws['CO'+str(i)].value = 0
        ws['CP'+str(i)].value = 0
        
    wb.save(book)
    return("så er der wiped")
#print(wipe("hs"))

def loopBadminton(): 
    lst = ["hs","ds","hd","md","dd"]
    tekstlst = [tekst[HSstart:HSslut],tekst[DSstart:DSslut],tekst[HDstart:HDslut],tekst[MDstart:MDslut],tekst[DDstart:DDslut]]
    for i in range(len(lst)):
        (wipe(lst[i]))
print(loopBadminton())
def excelDel(lst): 
    
    for i in range(len(lst)):
        wipe(lst[i])
       


#print(excelDel(["hs","hd","ds","dd","md"]))

driver.quit()
